from __future__ import annotations

from collections import Counter, defaultdict
from datetime import date, datetime, time
from pathlib import Path
import json
import math
import re
from typing import Any

from openpyxl import load_workbook


SOURCE_DIR = Path(r"D:\Projects\data\Hotels")
OUTPUT_TS = Path("src/db/realistic-hotel-fixtures.ts")
OUTPUT_DOC = Path("docs/real-hotel-fixture-mapping.md")
SNAPSHOT_DATE = "2026-05-25"

HOTEL_CONFIG = {
    "pecos": {
        "name_match": "Pecos",
        "id": "hotel_realistic_pecos",
        "name": "B&J Hotel and Apartments Pecos TX",
        "city": "Pecos",
        "state": "TX",
        "timezone": "America/Chicago",
        "room_rates": {
            "King": 8073,
            "Double Queen": 8775,
            "Suites": 16146,
        },
        "capacities": {
            "King": 2,
            "Double Queen": 4,
            "Suites": 4,
        },
    },
    "roswell": {
        "name_match": "Roswell",
        "id": "hotel_realistic_roswell",
        "name": "B&J Hotels Roswell NM",
        "city": "Roswell",
        "state": "NM",
        "timezone": "America/Denver",
        "room_rates": {
            "Studio": 7819,
            "Studio Plus": 8953,
            "One Bed One Bath": 10378,
            "Two Bed Two Bath": 15108,
        },
        "capacities": {
            "Studio": 2,
            "Studio Plus": 2,
            "One Bed One Bath": 3,
            "Two Bed Two Bath": 5,
        },
    },
}

FIRST_NAMES = [
    "Avery",
    "Blair",
    "Casey",
    "Dakota",
    "Emerson",
    "Finley",
    "Gray",
    "Harper",
    "Indigo",
    "Jordan",
    "Kai",
    "Logan",
    "Morgan",
    "Noel",
    "Oakley",
    "Parker",
    "Quinn",
    "Reese",
    "Sawyer",
    "Taylor",
    "Uma",
    "Val",
    "Wren",
    "Yael",
    "Zion",
]

LAST_NAMES = [
    "Adams",
    "Bennett",
    "Cole",
    "Diaz",
    "Ellis",
    "Foster",
    "Garcia",
    "Hayes",
    "Irwin",
    "Jensen",
    "Keller",
    "Lane",
    "Morris",
    "Nolan",
    "Owens",
    "Price",
    "Quintero",
    "Reed",
    "Stone",
    "Turner",
    "Underwood",
    "Vega",
    "Walker",
    "Young",
    "Zimmer",
]

SOURCE_CYCLE = ["direct", "ota", "phone", "web", "corporate", "walk-in"]
HOUSEKEEPING_STATUS_CYCLE = ["dirty", "cleaning", "inspection"]
MAINTENANCE_TEMPLATES = [
    ("HVAC not cooling", "high", "open", 0),
    ("Bathroom sink leak", "medium", "in-progress", 0),
    ("Door lock vendor hold", "critical", "blocked", 1),
    ("Loose towel bar reported by housekeeping", "medium", "pending-review", 0),
    ("Remote battery replaced", "low", "resolved", -1),
    ("Duplicate noise complaint", "low", "cancelled", -2),
]
BOOKING_REQUEST_TEMPLATES = [
    ("new", "King", 1, 2, "Needs a quiet room away from traffic."),
    ("contacted", "Double Queen", 1, 3, "Asked about weekly rate and late arrival."),
    ("accepted", "Studio Plus", 2, 5, "Accepted after phone follow-up."),
    ("declined", "Suite", 0, 1, "No matching room type at requested rate."),
]


def normalize_key(value: str) -> str:
    return re.sub(r"[^a-z0-9]+", "-", value.lower()).strip("-")


def clean(value: Any) -> Any:
    if value is None:
        return None
    if isinstance(value, str):
        value = value.strip()
        return value or None
    return value


def money_to_cents(value: Any) -> int:
    if value is None:
        return 0
    return int(round(float(value) * 100))


def floor_from_room_number(number: str, zone: str | None) -> int:
    if zone:
        zone_match = re.search(r"(\d+)", zone)
        if zone_match:
            return int(zone_match.group(1))
    room_match = re.search(r"\d", number)
    if room_match:
        return int(room_match.group(0))
    return 1


def map_condition(condition: Any) -> str:
    return normalize_key(str(condition or "clean"))


def map_occupancy(occupancy: Any) -> str:
    value = str(occupancy or "Vacant").lower()
    if "stayover" in value:
        return "occupied-stayover"
    if "departing" in value:
        return "occupied-departing"
    return "vacant"


def map_app_status(condition: str, occupancy_state: str, has_arrival_due: bool) -> str:
    if occupancy_state.startswith("occupied"):
        return "occupied"
    if condition == "dirty":
        return "dirty"
    if has_arrival_due:
        return "ready"
    return "available"


def as_date(value: Any) -> date | None:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    return None


def as_time_string(value: Any) -> str | None:
    if isinstance(value, datetime):
        return value.time().isoformat(timespec="seconds")
    if isinstance(value, time):
        return value.isoformat(timespec="seconds")
    if isinstance(value, str):
        match = re.match(r"^(\d{2}:\d{2}:\d{2})", value)
        if match:
            return match.group(1)
    return None


def payment_bucket(processed_time: str | None) -> str:
    if not processed_time:
        return "unknown"
    hour = int(processed_time[:2])
    if hour < 6:
        return "overnight"
    if hour < 11:
        return "morning"
    if hour < 15:
        return "midday"
    if hour < 19:
        return "evening-check-in"
    return "late-evening"


def find_file(hotel_key: str, report_name: str) -> Path:
    config = HOTEL_CONFIG[hotel_key]
    matches = [
        path
        for path in SOURCE_DIR.glob("*.xlsx")
        if config["name_match"] in path.name and report_name in path.name
    ]
    if len(matches) != 1:
        raise RuntimeError(f"Expected one {report_name} file for {hotel_key}, found {len(matches)}.")
    return matches[0]


def parse_housekeeping(hotel_key: str) -> tuple[list[dict[str, Any]], dict[str, Any]]:
    config = HOTEL_CONFIG[hotel_key]
    workbook = load_workbook(find_file(hotel_key, "HousekeepingReport"), data_only=True, read_only=True)
    worksheet = workbook["Detailed"]
    rows = list(worksheet.iter_rows(values_only=True))
    headers = [str(cell).strip() if cell is not None else "" for cell in rows[0]]
    column = {name: index for index, name in enumerate(headers)}
    rooms: list[dict[str, Any]] = []

    for index, row in enumerate(rows[1:], start=1):
        room_number = clean(row[column["Room #"]])
        if not room_number:
            continue
        room_number = str(room_number)
        room_type = str(clean(row[column["Room Class"]]) or "Standard")
        zone = clean(row[column["Zone"]])
        condition = map_condition(row[column["Condition"]])
        occupancy_state = map_occupancy(row[column["Occupancy"]])
        has_current_stay = bool(clean(row[column["Current Stay Guests"]]) or clean(row[column["Current Stay Guest Name"]]))
        has_arrival_due = bool(clean(row[column["Arrival Due Guests"]]) or clean(row[column["Arrival Due Guest Name"]]))
        room_id = f"{config['id']}_room_{normalize_key(room_number)}"
        rooms.append(
            {
                "id": room_id,
                "number": room_number,
                "roomType": room_type,
                "floor": floor_from_room_number(room_number, str(zone) if zone else None),
                "zone": zone or "Default Zone",
                "capacity": config["capacities"].get(room_type, 2),
                "nightlyRateCents": config["room_rates"].get(room_type, 8995),
                "currentAppStatus": map_app_status(condition, occupancy_state, has_arrival_due),
                "housekeepingCondition": condition,
                "occupancyState": occupancy_state,
                "hasCurrentStay": has_current_stay,
                "hasArrivalDue": has_arrival_due,
                "sortOrder": index,
            }
        )

    summary = {
        "roomsTotal": len(rooms),
        "roomTypeCounts": dict(Counter(room["roomType"] for room in rooms)),
        "zoneCounts": dict(Counter(room["zone"] for room in rooms)),
        "conditionCounts": dict(Counter(room["housekeepingCondition"] for room in rooms)),
        "occupancyCounts": dict(Counter(room["occupancyState"] for room in rooms)),
        "currentStayRooms": sum(1 for room in rooms if room["hasCurrentStay"]),
        "arrivalDueRooms": sum(1 for room in rooms if room["hasArrivalDue"]),
        "compoundRoomNumbers": sum(1 for room in rooms if "/" in room["number"]),
    }
    return rooms, summary


def parse_summary_sheet(hotel_key: str) -> dict[str, Any]:
    workbook = load_workbook(find_file(hotel_key, "TransactionsReport"), data_only=True, read_only=True)
    worksheet = workbook["Summary"]
    rows = [[clean(value) for value in row[:4]] for row in worksheet.iter_rows(values_only=True)]
    payment_method_totals: dict[str, int] = {}
    card_type_totals: dict[str, dict[str, int]] = {}
    card_transaction_totals: dict[str, dict[str, int]] = {}
    net_total_cents = 0
    section = "payment"

    for row in rows:
        label = row[0]
        if label is None:
            continue
        label_text = str(label)
        if label_text == "Credit Card Type":
            section = "card-type"
            continue
        if label_text == "Credit Card Transaction Type":
            section = "transaction-type"
            continue
        if section == "payment":
            if label_text == "Payment Method":
                continue
            if label_text == "Net Total ($)":
                net_total_cents = money_to_cents(row[1])
                continue
            payment_method_totals[label_text] = money_to_cents(row[1])
        elif section == "card-type":
            if label_text == "Credit Card Total ($)":
                card_type_totals["total"] = {
                    "capturedCents": money_to_cents(row[1]),
                    "refundedCents": money_to_cents(row[2]),
                    "netCents": money_to_cents(row[3]),
                }
                continue
            card_type_totals[label_text] = {
                "capturedCents": money_to_cents(row[1]),
                "refundedCents": money_to_cents(row[2]),
                "netCents": money_to_cents(row[3]),
            }
        elif section == "transaction-type":
            if label_text in {"Credit Card Transaction Type"}:
                continue
            card_transaction_totals[label_text] = {
                "count": int(row[1] or 0),
                "totalCents": money_to_cents(row[2]),
            }

    return {
        "netTotalCents": net_total_cents,
        "paymentMethodTotals": payment_method_totals,
        "cardTypeTotals": card_type_totals,
        "cardTransactionTotals": card_transaction_totals,
    }


def parse_payment_events(hotel_key: str) -> tuple[list[dict[str, Any]], list[dict[str, Any]], dict[str, Any]]:
    workbook = load_workbook(find_file(hotel_key, "TransactionsReport"), data_only=True, read_only=True)
    worksheet = workbook["Detailed"]
    account_map: dict[str, str] = {}
    business_date: date | None = None
    in_table = False
    events: list[dict[str, Any]] = []

    def account_ref(raw_value: Any) -> str | None:
        if raw_value is None:
            return None
        key = str(raw_value)
        if key not in account_map:
            account_map[key] = f"{hotel_key}-account-{len(account_map) + 1:03d}"
        return account_map[key]

    for row in worksheet.iter_rows(values_only=True):
        values = [clean(value) for value in row[:12]]
        maybe_date = as_date(values[0])
        if maybe_date:
            business_date = maybe_date
            in_table = False
            continue
        if str(values[1]).lower() == "reservation/ account name" and str(values[11]).lower() == "amount ($)":
            in_table = True
            continue
        if not in_table or business_date is None:
            continue
        if str(values[1]).lower() in {"payment method", "credit card type", "credit card transaction type"}:
            in_table = False
            continue
        amount = values[11]
        if not isinstance(amount, (int, float)):
            continue

        method = str(values[4] or "Unknown")
        card_transaction_type = None if values[6] in (None, "--") else str(values[6])
        card_type = None if values[7] in (None, "--") else str(values[7])
        card_method = None if values[9] in (None, "--") else str(values[9])
        revenue_eligible = card_transaction_type != "Decline"
        processed_time = as_time_string(values[3])
        event_index = len(events) + 1
        events.append(
            {
                "id": f"{hotel_key}-payment-{event_index:03d}",
                "businessDate": business_date.isoformat(),
                "processedTime": processed_time,
                "timeBucket": payment_bucket(processed_time),
                "method": method,
                "cardTransactionType": card_transaction_type,
                "cardNetwork": card_type,
                "cardMethod": card_method,
                "amountCents": money_to_cents(amount),
                "revenueEligible": revenue_eligible,
                "syntheticAccountRef": account_ref(values[2]),
            }
        )

    daily: dict[str, dict[str, Any]] = defaultdict(
        lambda: {
            "eventCount": 0,
            "revenueEventCount": 0,
            "declineCount": 0,
            "refundCount": 0,
            "grossPositiveCents": 0,
            "negativeCents": 0,
            "netRevenueCents": 0,
            "cashCents": 0,
            "checkCents": 0,
            "creditCardCents": 0,
        }
    )
    method_counts = Counter()
    time_bucket_counts = Counter()
    repeated_accounts = Counter()
    positive_amounts: list[int] = []

    for event in events:
        day = daily[event["businessDate"]]
        amount_cents = event["amountCents"]
        day["eventCount"] += 1
        method_counts[event["method"]] += 1
        time_bucket_counts[event["timeBucket"]] += 1
        if event["syntheticAccountRef"]:
            repeated_accounts[event["syntheticAccountRef"]] += 1
        if amount_cents > 0:
            day["grossPositiveCents"] += amount_cents
            positive_amounts.append(amount_cents)
        if amount_cents < 0:
            day["negativeCents"] += amount_cents
        if event["cardTransactionType"] == "Decline":
            day["declineCount"] += 1
        if event["cardTransactionType"] == "Refund":
            day["refundCount"] += 1
        if event["revenueEligible"]:
            day["revenueEventCount"] += 1
            day["netRevenueCents"] += amount_cents
            if event["method"] == "Cash":
                day["cashCents"] += amount_cents
            elif event["method"] == "Check":
                day["checkCents"] += amount_cents
            elif event["method"] == "Credit Card":
                day["creditCardCents"] += amount_cents

    daily_summaries = [
        {"businessDate": business_date, **values}
        for business_date, values in sorted(daily.items())
    ]
    stats = {
        "eventCount": len(events),
        "activeDays": len(daily_summaries),
        "methodCounts": dict(method_counts),
        "timeBucketCounts": dict(time_bucket_counts),
        "repeatedSyntheticAccounts": sum(1 for count in repeated_accounts.values() if count > 1),
        "maxEventsForOneSyntheticAccount": max(repeated_accounts.values(), default=0),
        "positiveAmountDistributionCents": distribution(positive_amounts),
    }
    return events, daily_summaries, stats


def distribution(values: list[int]) -> dict[str, int]:
    sorted_values = sorted(values)
    if not sorted_values:
        return {}

    def percentile(pct: float) -> int:
        index = (len(sorted_values) - 1) * pct
        low = math.floor(index)
        high = math.ceil(index)
        if low == high:
            return sorted_values[low]
        return round(sorted_values[low] * (high - index) + sorted_values[high] * (index - low))

    return {
        "min": sorted_values[0],
        "p25": percentile(0.25),
        "median": percentile(0.5),
        "p75": percentile(0.75),
        "max": sorted_values[-1],
    }


def synthetic_guest_name(index: int) -> str:
    first = FIRST_NAMES[index % len(FIRST_NAMES)]
    last = LAST_NAMES[(index * 7) % len(LAST_NAMES)]
    return f"{first} {last}"


def build_reservations_and_guests(hotel_key: str, rooms: list[dict[str, Any]]) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    config = HOTEL_CONFIG[hotel_key]
    guests: list[dict[str, Any]] = []
    reservations: list[dict[str, Any]] = []
    eligible_rooms = [room for room in rooms if room["hasCurrentStay"] or room["hasArrivalDue"]]

    for index, room in enumerate(eligible_rooms, start=1):
        guest_id = f"{config['id']}_guest_{index:03d}"
        reservation_id = f"{config['id']}_reservation_{index:03d}"
        name = synthetic_guest_name(index)
        phone_suffix = 1000 + index
        is_arrival = bool(room["hasArrivalDue"])
        is_departing = room["occupancyState"] == "occupied-departing"
        check_in_offset = 0 if is_arrival else -((index % 9) + 1)
        check_out_offset = (index % 5) + 1
        if is_departing:
            check_out_offset = 0
        adults = min(max(1, room["capacity"] - 1), room["capacity"])
        children = 1 if room["capacity"] >= 4 and index % 3 == 0 else 0
        nights = max(1, check_out_offset - check_in_offset)
        status = "confirmed" if is_arrival else "checked-in"
        if is_arrival and index % 4 == 0:
            status = "pending"
        guests.append(
            {
                "id": guest_id,
                "fullName": name,
                "email": f"guest{index:03d}.{hotel_key}@example.invalid",
                "phone": f"555-{phone_suffix:04d}",
                "notes": "",
            }
        )
        reservations.append(
            {
                "id": reservation_id,
                "guestId": guest_id,
                "roomId": room["id"],
                "roomNumber": room["number"],
                "checkInOffsetDays": check_in_offset,
                "checkOutOffsetDays": check_out_offset,
                "adults": adults,
                "children": children,
                "nightlyRateCents": room["nightlyRateCents"],
                "totalCents": nights * room["nightlyRateCents"],
                "source": SOURCE_CYCLE[index % len(SOURCE_CYCLE)],
                "status": status,
                "operationalState": "arrival-due" if is_arrival else room["occupancyState"],
                "notes": "Synthetic fixture generated from anonymized operational shape.",
            }
        )
    return guests, reservations


def build_housekeeping_tasks(hotel_key: str, rooms: list[dict[str, Any]]) -> list[dict[str, Any]]:
    config = HOTEL_CONFIG[hotel_key]
    tasks: list[dict[str, Any]] = []
    for room in rooms:
        needs_task = room["housekeepingCondition"] == "dirty" or room["occupancyState"] == "occupied-departing"
        if not needs_task:
            continue
        title = "Departure turn" if room["occupancyState"] == "occupied-departing" else "Clean room"
        if room["hasArrivalDue"]:
            title = "Priority arrival clean"
        if room["occupancyState"] == "occupied-stayover":
            title = "Stayover refresh"
        task_number = len(tasks) + 1
        status = HOUSEKEEPING_STATUS_CYCLE[(task_number - 1) % len(HOUSEKEEPING_STATUS_CYCLE)]
        room["currentAppStatus"] = "cleaning" if status in {"cleaning", "inspection"} else "dirty"
        tasks.append(
            {
                "id": f"{config['id']}_hk_{task_number:03d}",
                "roomId": room["id"],
                "roomNumber": room["number"],
                "title": title,
                "status": status,
                "dueOffsetDays": 0,
                "notes": "Generated from source housekeeping condition and occupancy snapshot.",
            }
        )
    assigned_room_ids = {task["roomId"] for task in tasks}
    fill_candidates = [
        room
        for room in rooms
        if room["id"] not in assigned_room_ids and room["currentAppStatus"] in {"available", "ready"} and not room["hasCurrentStay"]
    ]
    while len([task for task in tasks if task["status"] != "blocked"]) < 4 and fill_candidates:
        room = fill_candidates.pop(0)
        task_number = len(tasks) + 1
        status = HOUSEKEEPING_STATUS_CYCLE[(task_number - 1) % len(HOUSEKEEPING_STATUS_CYCLE)]
        room["currentAppStatus"] = "cleaning" if status in {"cleaning", "inspection"} else "dirty"
        tasks.append(
            {
                "id": f"{config['id']}_hk_sample_{task_number:03d}",
                "roomId": room["id"],
                "roomNumber": room["number"],
                "title": "Sample room readiness task",
                "status": status,
                "dueOffsetDays": 0,
                "notes": "Synthetic task added so every housekeeping demo user has an active workflow.",
            }
        )
    return tasks


def candidate_service_rooms(rooms: list[dict[str, Any]]) -> list[dict[str, Any]]:
    candidates = [
        room
        for room in rooms
        if room["currentAppStatus"] in {"available", "ready", "dirty"} and not room["hasCurrentStay"]
    ]
    return candidates or [room for room in rooms if not room["hasCurrentStay"]] or rooms


def build_maintenance_tickets(hotel_key: str, rooms: list[dict[str, Any]], housekeeping_tasks: list[dict[str, Any]]) -> list[dict[str, Any]]:
    config = HOTEL_CONFIG[hotel_key]
    candidates = candidate_service_rooms(rooms)
    tickets: list[dict[str, Any]] = []
    for index, (title, priority, status, due_offset_days) in enumerate(MAINTENANCE_TEMPLATES, start=1):
        room = candidates[((index - 1) * 3) % len(candidates)]
        ticket_id = f"{config['id']}_mt_{index:03d}"
        if status in {"open", "in-progress", "blocked"}:
            room["currentAppStatus"] = "maintenance"
        if status == "pending-review":
            room["currentAppStatus"] = "dirty"
            housekeeping_tasks.append(
                {
                    "id": f"{config['id']}_hk_issue_{index:03d}",
                    "roomId": room["id"],
                    "roomNumber": room["number"],
                    "title": "Hold for maintenance review",
                    "status": "blocked",
                    "dueOffsetDays": 0,
                    "notes": title,
                }
            )
        tickets.append(
            {
                "id": ticket_id,
                "roomId": room["id"],
                "roomNumber": room["number"],
                "title": title,
                "priority": priority,
                "status": status,
                "dueOffsetDays": due_offset_days,
            }
        )
    return tickets


def preferred_request_room_type(hotel_key: str, requested_type: str) -> str:
    configured_types = HOTEL_CONFIG[hotel_key]["room_rates"].keys()
    if requested_type in configured_types:
        return requested_type
    return next(iter(configured_types))


def build_booking_requests(hotel_key: str) -> list[dict[str, Any]]:
    config = HOTEL_CONFIG[hotel_key]
    requests: list[dict[str, Any]] = []
    for index, (status, requested_type, check_in_offset, check_out_offset, message) in enumerate(BOOKING_REQUEST_TEMPLATES, start=1):
        name = synthetic_guest_name(index + 80)
        requests.append(
            {
                "id": f"{config['id']}_request_{index:03d}",
                "fullName": name,
                "email": f"request{index:03d}.{hotel_key}@example.invalid",
                "phone": f"555-{3000 + index:04d}",
                "checkInOffsetDays": check_in_offset,
                "checkOutOffsetDays": check_out_offset,
                "requestedRoomType": preferred_request_room_type(hotel_key, requested_type),
                "status": status,
                "message": message,
            }
        )
    return requests


def build_fixture(hotel_key: str) -> dict[str, Any]:
    config = HOTEL_CONFIG[hotel_key]
    rooms, housekeeping_summary = parse_housekeeping(hotel_key)
    guests, reservations = build_reservations_and_guests(hotel_key, rooms)
    housekeeping_tasks = build_housekeeping_tasks(hotel_key, rooms)
    maintenance_tickets = build_maintenance_tickets(hotel_key, rooms, housekeeping_tasks)
    booking_requests = build_booking_requests(hotel_key)
    payment_summary = parse_summary_sheet(hotel_key)
    payment_events, payment_daily_summaries, payment_stats = parse_payment_events(hotel_key)
    return {
        "hotel": {
            "id": config["id"],
            "name": config["name"],
            "city": config["city"],
            "state": config["state"],
            "timezone": config["timezone"],
            "checkInTime": "15:00",
            "checkOutTime": "11:00",
        },
        "source": {
            "snapshotDate": SNAPSHOT_DATE,
            "sourceDirectory": str(SOURCE_DIR),
            "containsRealGuestData": False,
            "notes": [
                "Guest, account, staff, and card-identifying fields are not copied from source workbooks.",
                "Room inventory, room classes, housekeeping condition, occupancy state, payment timing, and payment amounts preserve the source operational shape.",
            ],
        },
        "housekeepingSummary": housekeeping_summary,
        "paymentSummary": {**payment_summary, **payment_stats},
        "rooms": rooms,
        "guests": guests,
        "reservations": reservations,
        "housekeepingTasks": housekeeping_tasks,
        "maintenanceTickets": maintenance_tickets,
        "bookingRequests": booking_requests,
        "paymentDailySummaries": payment_daily_summaries,
        "paymentEvents": payment_events,
    }


def write_typescript(fixtures: list[dict[str, Any]]) -> None:
    OUTPUT_TS.parent.mkdir(parents=True, exist_ok=True)
    fixture_json = json.dumps(fixtures, indent=2)
    content = f'''import type {{ BookingRequestStatus, MaintenancePriority, MaintenanceStatus, ReservationStatus, RoomStatus }} from "@/lib/types";

export type HousekeepingCondition = "clean" | "dirty";
export type OccupancyState = "vacant" | "occupied-stayover" | "occupied-departing";
export type PaymentMethod = "Cash" | "Check" | "Credit Card" | "Unknown";
export type CardTransactionType = "Capture" | "Decline" | "Refund";
export type TimeBucket = "overnight" | "morning" | "midday" | "evening-check-in" | "late-evening" | "unknown";

export type RealisticHotelRoomFixture = {{
  id: string;
  number: string;
  roomType: string;
  floor: number;
  zone: string;
  capacity: number;
  nightlyRateCents: number;
  currentAppStatus: RoomStatus;
  housekeepingCondition: HousekeepingCondition;
  occupancyState: OccupancyState;
  hasCurrentStay: boolean;
  hasArrivalDue: boolean;
  sortOrder: number;
}};

export type RealisticHotelGuestFixture = {{
  id: string;
  fullName: string;
  email: string;
  phone: string;
  notes: string;
}};

export type RealisticHotelReservationFixture = {{
  id: string;
  guestId: string;
  roomId: string;
  roomNumber: string;
  checkInOffsetDays: number;
  checkOutOffsetDays: number;
  adults: number;
  children: number;
  nightlyRateCents: number;
  totalCents: number;
  source: string;
  status: ReservationStatus;
  operationalState: OccupancyState | "arrival-due";
  notes: string;
}};

export type RealisticHousekeepingTaskFixture = {{
  id: string;
  roomId: string;
  roomNumber: string;
  title: string;
  status: "dirty" | "cleaning" | "inspection" | "blocked";
  dueOffsetDays: number;
  notes: string;
}};

export type RealisticMaintenanceTicketFixture = {{
  id: string;
  roomId: string;
  roomNumber: string;
  title: string;
  priority: MaintenancePriority;
  status: MaintenanceStatus;
  dueOffsetDays: number;
}};

export type RealisticBookingRequestFixture = {{
  id: string;
  fullName: string;
  email: string;
  phone: string;
  checkInOffsetDays: number;
  checkOutOffsetDays: number;
  requestedRoomType: string;
  status: BookingRequestStatus;
  message: string;
}};

export type PaymentDailySummaryFixture = {{
  businessDate: string;
  eventCount: number;
  revenueEventCount: number;
  declineCount: number;
  refundCount: number;
  grossPositiveCents: number;
  negativeCents: number;
  netRevenueCents: number;
  cashCents: number;
  checkCents: number;
  creditCardCents: number;
}};

export type PaymentEventFixture = {{
  id: string;
  businessDate: string;
  processedTime: string | null;
  timeBucket: TimeBucket;
  method: PaymentMethod;
  cardTransactionType: CardTransactionType | null;
  cardNetwork: "Visa" | "MC" | "Amex" | "Discover" | null;
  cardMethod: "virtual" | null;
  amountCents: number;
  revenueEligible: boolean;
  syntheticAccountRef: string | null;
}};

export type RealisticHotelFixture = {{
  hotel: {{
    id: string;
    name: string;
    city: string;
    state: string;
    timezone: string;
    checkInTime: string;
    checkOutTime: string;
  }};
  source: {{
    snapshotDate: string;
    sourceDirectory: string;
    containsRealGuestData: false;
    notes: string[];
  }};
  housekeepingSummary: {{
    roomsTotal: number;
    roomTypeCounts: Record<string, number>;
    zoneCounts: Record<string, number>;
    conditionCounts: Record<string, number>;
    occupancyCounts: Record<string, number>;
    currentStayRooms: number;
    arrivalDueRooms: number;
    compoundRoomNumbers: number;
  }};
  paymentSummary: {{
    netTotalCents: number;
    paymentMethodTotals: Record<string, number>;
    cardTypeTotals: Record<string, {{ capturedCents: number; refundedCents: number; netCents: number }}>;
    cardTransactionTotals: Record<string, {{ count: number; totalCents: number }}>;
    eventCount: number;
    activeDays: number;
    methodCounts: Record<string, number>;
    timeBucketCounts: Record<string, number>;
    repeatedSyntheticAccounts: number;
    maxEventsForOneSyntheticAccount: number;
    positiveAmountDistributionCents: Record<string, number>;
  }};
  rooms: RealisticHotelRoomFixture[];
  guests: RealisticHotelGuestFixture[];
  reservations: RealisticHotelReservationFixture[];
  housekeepingTasks: RealisticHousekeepingTaskFixture[];
  maintenanceTickets: RealisticMaintenanceTicketFixture[];
  bookingRequests: RealisticBookingRequestFixture[];
  paymentDailySummaries: PaymentDailySummaryFixture[];
  paymentEvents: PaymentEventFixture[];
}};

// Generated by scripts/organize-real-hotel-data.py from anonymized operational fields in the hotel exports.
// Do not add real guest names, reservation/account IDs, staff processor names, card last-four values, or notes here.
export const realisticHotelFixtures: RealisticHotelFixture[] = {fixture_json};
'''
    OUTPUT_TS.write_text(content, encoding="utf-8", newline="\n")


def write_mapping_doc(fixtures: list[dict[str, Any]]) -> None:
    OUTPUT_DOC.parent.mkdir(parents=True, exist_ok=True)
    lines = [
        "# Real Hotel Fixture Mapping",
        "",
        "Generated by `scripts/organize-real-hotel-data.py` from the Excel exports in `D:\\Projects\\data\\Hotels`.",
        "",
        "The generated TypeScript fixture is:",
        "",
        "- `src/db/realistic-hotel-fixtures.ts`",
        "",
        "The fixture is intentionally PII-safe. It does not copy real guest names, card last-four values, reservation/account numbers, staff processor names, emails, phone numbers, or free-text guest notes.",
        "",
        "## Implementation Status",
        "",
        "- Demo mode loads the realistic Pecos and Roswell hotels from this fixture.",
        "- Demo login users point at the Pecos fixture hotel, with the owner user also seeing Roswell in the portfolio.",
        "- `src/db/seed.ts` uses this fixture for hosted database seeding.",
        "- Sample booking requests, housekeeping states, and maintenance tickets are synthetic additions for full app testing.",
        "- Payment exports are organized in the fixture, but not persisted by `src/db/seed.ts` yet because the current schema has no payment transaction table.",
        "",
        "## Current App Mapping",
        "",
        "- `hotel` maps to `hotels` and `hotel_settings`.",
        "- `rooms[].currentAppStatus` maps to the current `rooms.status` column.",
        "- `rooms[].housekeepingCondition` and `rooms[].occupancyState` are supplemental fields that the current schema cannot store separately yet.",
        "- `guests` and `reservations` are synthetic, generated from current-stay and arrival-due counts.",
        "- `reservations[].checkInOffsetDays` and `checkOutOffsetDays` are relative offsets for dynamic seed dates.",
        "- `housekeepingTasks` include dirty, cleaning, inspection, and blocked sample states.",
        "- `maintenanceTickets` include open, in-progress, blocked, pending-review, resolved, and cancelled statuses.",
        "- `bookingRequests` include new, contacted, accepted, and declined statuses.",
        "- `paymentDailySummaries` and `paymentEvents` should wait for a payment table or demo-only payment UI. Current revenue should not treat declined payments as revenue.",
        "",
        "## Fixture Counts",
        "",
    ]
    for fixture in fixtures:
        lines.extend(
            [
                f"### {fixture['hotel']['name']}",
                "",
                f"- Rooms: {len(fixture['rooms'])}",
                f"- Synthetic guests: {len(fixture['guests'])}",
                f"- Synthetic reservations: {len(fixture['reservations'])}",
                f"- Housekeeping tasks: {len(fixture['housekeepingTasks'])}",
                f"- Maintenance tickets: {len(fixture['maintenanceTickets'])}",
                f"- Booking requests: {len(fixture['bookingRequests'])}",
                f"- Payment daily summaries: {len(fixture['paymentDailySummaries'])}",
                f"- Payment events: {len(fixture['paymentEvents'])}",
                f"- Revenue summary excluding declines: ${fixture['paymentSummary']['netTotalCents'] / 100:,.2f}",
                "",
            ]
        )
    lines.extend(
        [
            "## Remaining Integration Order",
            "",
            "1. Add schema fields for separated room condition and occupancy state before trying to persist `housekeepingCondition` and `occupancyState`.",
            "2. Add a payment transaction table before persisting `paymentEvents`.",
            "3. Compute revenue from payment events with `revenueEligible === true`; declined payment attempts should be counted as failed payments, not revenue.",
            "",
        ]
    )
    OUTPUT_DOC.write_text("\n".join(lines), encoding="utf-8", newline="\n")


def main() -> None:
    fixtures = [build_fixture("pecos"), build_fixture("roswell")]
    write_typescript(fixtures)
    write_mapping_doc(fixtures)
    print(f"Wrote {OUTPUT_TS}")
    print(f"Wrote {OUTPUT_DOC}")


if __name__ == "__main__":
    main()
